import openpyxl

from mysql2doc.MySql import MySql

workbook = openpyxl.load_workbook("batch.xlsx")
worksheet = workbook["字段信息采集表"]

mysql = None
try:
    mysql = MySql()
    row = 3
    for _, table in enumerate(mysql.generateTableData()):
        for field in table.fields:
            # 序号
            worksheet.cell(row=row, column=1, value=row - 2)
            # 实体名（ * ）
            worksheet.cell(row=row, column=2, value=table.comment)
            # 中文逻辑含义（ * ）
            worksheet.cell(row=row, column=3, value=table.comment + '-' + field.comment)
            # 用语结构信息（ * ）
            worksheet.cell(row=row, column=4, value=table.name + field.name)
            # 域名（ * ）
            worksheet.cell(row=row, column=5, value=field.type)
            # 数据类型（ * ）
            worksheet.cell(row=row, column=6, value=field.type)
            # 长度
            worksheet.cell(row=row, column=7, value=field.type)
            # 小数点
            worksheet.cell(row=row, column=8, value=field.type)
            # 是否主键（ * ）
            worksheet.cell(row=row, column=9, value=field.isPK)
            # 是否Not Null（ * ）
            worksheet.cell(row=row, column=10, value=field.nullable)
            # 实际数据类型
            worksheet.cell(row=row, column=11, value="")
            # 实际长度
            worksheet.cell(row=row, column=12, value="")
            # 实际小数点
            worksheet.cell(row=row, column=13, value="")
            # 字段物理名（ * ）
            worksheet.cell(row=row, column=14, value="")
            # 字段表示名（ * ）
            worksheet.cell(row=row, column=15, value="")
            # 中文表名（ * ）
            worksheet.cell(row=row, column=16, value="")
            # 物理表名（ * ）
            worksheet.cell(row=row, column=17, value="")
            # 说明（ * ）
            worksheet.cell(row=row, column=18, value=field.comment)

            row = row + 1
finally:
    MySql.close(mysql)

workbook.save("result.xlsx")
